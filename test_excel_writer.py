#!/usr/bin/env python3
"""
Test Script for Excel Writer Functions
======================================

This script tests the Excel writer functions to ensure:
1. Work Name row has no fill (cell.fill.patternType is None)
2. Correct allocation codes appear for subcategories 'M&P', 'RSP', and default (None/other)
3. Mocks workbook writes to avoid file I/O where possible
"""

import unittest
import sys
import os
from unittest.mock import MagicMock, patch, Mock
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Add the project root to Python path
sys.path.insert(0, os.path.abspath('.'))

from features.estimates.writer import write_work_name_row, write_summary_section


class TestExcelWriter(unittest.TestCase):

    def setUp(self):
        """Set up a mock workbook and worksheet for each test."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def test_write_work_name_row_no_fill(self):
        """Test that the 'Work Name' row has no fill pattern."""
        work_description_data = {'description': 'Test Work Description'}
        current_row = 1
        
        # Call the function to write the work name row
        next_row = write_work_name_row(self.worksheet, work_description_data, current_row)
        
        # Get the cell that was written to
        cell = self.worksheet.cell(row=current_row, column=1)
        
        # Assert that the fill pattern is None (default)
        self.assertIsNone(cell.fill.patternType, "The 'Work Name' cell should have no fill pattern.")
        
        # Verify the cell value was set correctly
        expected_value = "Name Of Work : Test Work Description"
        self.assertEqual(cell.value, expected_value, "Cell value should match expected format.")
        
        # Verify function returns next row
        self.assertEqual(next_row, current_row + 1, "Function should return the next row number.")

    def test_write_work_name_row_with_empty_description(self):
        """Test work name row when description is empty."""
        work_description_data = {'description': ''}
        current_row = 2
        
        # Call the function
        write_work_name_row(self.worksheet, work_description_data, current_row)
        
        # Get the cell
        cell = self.worksheet.cell(row=current_row, column=1)
        
        # Verify no fill
        self.assertIsNone(cell.fill.patternType, "Cell should have no fill even with empty description.")
        
        # Verify the cell value handles empty description
        expected_value = "Name Of Work : "
        self.assertEqual(cell.value, expected_value)

    def test_write_work_name_row_with_missing_description(self):
        """Test work name row when description key is missing."""
        work_description_data = {}  # No description key
        current_row = 3
        
        # Call the function
        write_work_name_row(self.worksheet, work_description_data, current_row)
        
        # Get the cell
        cell = self.worksheet.cell(row=current_row, column=1)
        
        # Verify no fill
        self.assertIsNone(cell.fill.patternType, "Cell should have no fill even with missing description.")
        
        # Verify the cell value handles missing description
        expected_value = "Name Of Work : "
        self.assertEqual(cell.value, expected_value)

    @patch('features.estimates.writer.get_column_letter')
    def test_allocation_code_for_mp_subcategory(self, mock_get_column_letter):
        """Test that the correct allocation code is set for M&P subcategory."""
        # Mock get_column_letter to return simple column letters
        mock_get_column_letter.side_effect = lambda x: chr(ord('A') + x - 1)
        
        # Mock worksheet.cell to track calls
        mock_cells = {}
        original_cell = self.worksheet.cell
        
        def mock_cell_method(row, column, value=None):
            cell = original_cell(row, column, value)
            mock_cells[(row, column)] = cell
            return cell
        
        self.worksheet.cell = mock_cell_method
        
        # Call the function with M&P subcategory
        write_summary_section(self.worksheet, 2, 5, work_subcategory='M&P')
        
        # Find the allocation cell by searching for cells with 'Allocation' in the value
        allocation_cell = None
        for (row, col), cell in mock_cells.items():
            if cell.value and 'Allocation' in str(cell.value):
                allocation_cell = cell
                break
        
        self.assertIsNotNone(allocation_cell, "Allocation cell should be written.")
        self.assertIn('47000 050 443 32', allocation_cell.value, 
                     "M&P subcategory should use allocation code '47000 050 443 32'.")

    @patch('features.estimates.writer.get_column_letter')
    def test_allocation_code_for_rsp_subcategory(self, mock_get_column_letter):
        """Test that the correct allocation code is set for RSP subcategory."""
        # Mock get_column_letter to return simple column letters
        mock_get_column_letter.side_effect = lambda x: chr(ord('A') + x - 1)
        
        # Mock worksheet.cell to track calls
        mock_cells = {}
        original_cell = self.worksheet.cell
        
        def mock_cell_method(row, column, value=None):
            cell = original_cell(row, column, value)
            mock_cells[(row, column)] = cell
            return cell
        
        self.worksheet.cell = mock_cell_method
        
        # Call the function with RSP subcategory
        write_summary_section(self.worksheet, 2, 5, work_subcategory='RSP')
        
        # Find the allocation cell
        allocation_cell = None
        for (row, col), cell in mock_cells.items():
            if cell.value and 'Allocation' in str(cell.value):
                allocation_cell = cell
                break
        
        self.assertIsNotNone(allocation_cell, "Allocation cell should be written.")
        self.assertIn('47000 070 443 32', allocation_cell.value, 
                     "RSP subcategory should use allocation code '47000 070 443 32'.")

    @patch('features.estimates.writer.get_column_letter')
    def test_allocation_code_for_default_none_subcategory(self, mock_get_column_letter):
        """Test that the correct allocation code is set for default (None) subcategory."""
        # Mock get_column_letter to return simple column letters
        mock_get_column_letter.side_effect = lambda x: chr(ord('A') + x - 1)
        
        # Mock worksheet.cell to track calls
        mock_cells = {}
        original_cell = self.worksheet.cell
        
        def mock_cell_method(row, column, value=None):
            cell = original_cell(row, column, value)
            mock_cells[(row, column)] = cell
            return cell
        
        self.worksheet.cell = mock_cell_method
        
        # Call the function with None subcategory (default)
        write_summary_section(self.worksheet, 2, 5, work_subcategory=None)
        
        # Find the allocation cell
        allocation_cell = None
        for (row, col), cell in mock_cells.items():
            if cell.value and 'Allocation' in str(cell.value):
                allocation_cell = cell
                break
        
        self.assertIsNotNone(allocation_cell, "Allocation cell should be written.")
        self.assertIn('47000 070 443 32', allocation_cell.value, 
                     "Default (None) subcategory should use allocation code '47000 070 443 32'.")

    @patch('features.estimates.writer.get_column_letter')
    def test_allocation_code_for_unknown_subcategory(self, mock_get_column_letter):
        """Test that the correct allocation code is set for unknown subcategory."""
        # Mock get_column_letter to return simple column letters
        mock_get_column_letter.side_effect = lambda x: chr(ord('A') + x - 1)
        
        # Mock worksheet.cell to track calls
        mock_cells = {}
        original_cell = self.worksheet.cell
        
        def mock_cell_method(row, column, value=None):
            cell = original_cell(row, column, value)
            mock_cells[(row, column)] = cell
            return cell
        
        self.worksheet.cell = mock_cell_method
        
        # Call the function with an unknown subcategory
        write_summary_section(self.worksheet, 2, 5, work_subcategory='UNKNOWN')
        
        # Find the allocation cell
        allocation_cell = None
        for (row, col), cell in mock_cells.items():
            if cell.value and 'Allocation' in str(cell.value):
                allocation_cell = cell
                break
        
        self.assertIsNotNone(allocation_cell, "Allocation cell should be written.")
        self.assertIn('47000 070 443 32', allocation_cell.value, 
                     "Unknown subcategory should use default allocation code '47000 070 443 32'.")

    @patch('features.estimates.writer.get_column_letter')
    def test_summary_section_mock_workbook_writes(self, mock_get_column_letter):
        """Test that summary section writes are properly mocked to avoid file I/O."""
        # Mock get_column_letter to return simple column letters
        mock_get_column_letter.side_effect = lambda x: chr(ord('A') + x - 1)
        
        # Create a fully mocked worksheet
        mock_worksheet = MagicMock()
        mock_cell = MagicMock()
        mock_worksheet.cell.return_value = mock_cell
        mock_worksheet.append = MagicMock()
        mock_worksheet.merge_cells = MagicMock()
        
        # Call the function with mocked worksheet
        result = write_summary_section(mock_worksheet, 2, 5, work_subcategory='M&P')
        
        # Verify that worksheet methods were called (but no actual file I/O occurred)
        self.assertTrue(mock_worksheet.cell.called, "Worksheet cell method should be called.")
        self.assertTrue(mock_worksheet.append.called, "Worksheet append method should be called.")
        self.assertTrue(mock_worksheet.merge_cells.called, "Worksheet merge_cells method should be called.")
        
        # Verify the function returns the expected tuple
        self.assertIsInstance(result, tuple, "Function should return a tuple.")
        self.assertEqual(len(result), 2, "Function should return a tuple of 2 elements.")

    def test_work_name_row_font_and_alignment(self):
        """Test that the work name row has correct font and alignment properties."""
        work_description_data = {'description': 'Test Work'}
        current_row = 1
        
        write_work_name_row(self.worksheet, work_description_data, current_row)
        
        cell = self.worksheet.cell(row=current_row, column=1)
        
        # Test font properties
        self.assertTrue(cell.font.bold, "Work name cell should have bold font.")
        
        # Test alignment properties
        self.assertEqual(cell.alignment.horizontal, 'center', "Work name cell should be center-aligned horizontally.")
        self.assertEqual(cell.alignment.vertical, 'center', "Work name cell should be center-aligned vertically.")


def run_tests():
    """Run all tests and provide a summary."""
    print("🧪 Running Excel Writer Tests...")
    print("=" * 50)
    
    # Discover and run tests
    test_suite = unittest.TestLoader().loadTestsFromTestCase(TestExcelWriter)
    test_runner = unittest.TextTestRunner(verbosity=2)
    result = test_runner.run(test_suite)
    
    print("\n" + "=" * 50)
    print("📊 TEST SUMMARY:")
    print(f"   Tests run: {result.testsRun}")
    print(f"   Failures: {len(result.failures)}")
    print(f"   Errors: {len(result.errors)}")
    
    if result.failures:
        print("\n❌ FAILURES:")
        for test, traceback in result.failures:
            print(f"   - {test}: {traceback}")
    
    if result.errors:
        print("\n💥 ERRORS:")
        for test, traceback in result.errors:
            print(f"   - {test}: {traceback}")
    
    if result.wasSuccessful():
        print("\n✅ All tests passed!")
        return True
    else:
        print("\n❌ Some tests failed!")
        return False


if __name__ == '__main__':
    success = run_tests()
    sys.exit(0 if success else 1)
