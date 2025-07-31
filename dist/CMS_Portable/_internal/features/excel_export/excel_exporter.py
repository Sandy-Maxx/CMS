import pandas as pd
from datetime import datetime

def export_variation_report(work_details, schedule_items, firm_rates_by_item, updated_quantities, output_path):
    try:
        data = []
        for item in schedule_items:
            item_id = item['item_id']
            original_qty = item['quantity']
            new_qty = float(updated_quantities.get(str(item_id), original_qty))
            firm_rates = firm_rates_by_item.get(item_id, [])
            for rate in firm_rates:
                firm_name = rate['firm_name']
                unit_rate = rate['unit_rate']
                original_cost = original_qty * unit_rate
                new_cost = new_qty * unit_rate
                data.append({
                    'Item ID': item_id,
                    'Item Name': item['item_name'],
                    'Unit': item['unit'],
                    'Original Quantity': original_qty,
                    'New Quantity': new_qty,
                    'Firm Name': firm_name,
                    'Unit Rate': unit_rate,
                    'Original Cost': original_cost,
                    'New Cost': new_cost,
                    'Variation': new_cost - original_cost
                })
        df = pd.DataFrame(data)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vitiation Report', index=False)
            worksheet = writer.sheets['Vitiation Report']
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
        return True, "Report generated successfully"
    except Exception as e:
        return False, f"Error generating report: {str(e)}"

def export_work_to_excel(work_details, schedule_items, firm_rates_by_item, output_path):
    try:
        data = []
        for item in schedule_items:
            item_id = item['item_id']
            quantity = item['quantity']
            firm_rates = firm_rates_by_item.get(item_id, [])
            for rate in firm_rates:
                firm_name = rate['firm_name']
                unit_rate = rate['unit_rate']
                total_cost = quantity * unit_rate
                data.append({
                    'Work Name': work_details['work_name'],
                    'Item ID': item_id,
                    'Item Name': item['item_name'],
                    'Unit': item['unit'],
                    'Quantity': quantity,
                    'Firm Name': firm_name,
                    'Unit Rate': unit_rate,
                    'Total Cost': total_cost
                })
        df = pd.DataFrame(data)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Work Details', index=False)
            worksheet = writer.sheets['Work Details']
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
        return True, "Work details exported successfully"
    except Exception as e:
        return False, f"Error exporting work details: {str(e)}"

def export_estimate_to_excel(work_details, schedule_items_data, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from features.estimates.constants import COLUMN_HEADERS, COLUMN_MAPPING, GST_RATE

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Estimate Report"

        # Define styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1

        # Write the actual column headers
        header_start_row = current_row
        for col_idx, header_text in enumerate(COLUMN_HEADERS, 1):
            cell = worksheet.cell(row=header_start_row, column=col_idx, value=header_text)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1 # Move to the next row after headers

        # Write schedule data rows
        data_start_row = current_row
        for item in schedule_items_data:
            row_values = []
            for header in COLUMN_HEADERS:
                key = COLUMN_MAPPING.get(header)
                if key == "labour_amount":
                    qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                    labour_rate_col_idx = COLUMN_HEADERS.index("Labour rate in Rs.") + 1
                    formula = f"={get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(labour_rate_col_idx)}{current_row}"
                    row_values.append(formula)
                elif key == "total_in_rs":
                    qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                    rate_col_idx = COLUMN_HEADERS.index("Rate in Rs") + 1
                    labour_amount_col_idx = COLUMN_HEADERS.index("Labour Amount") + 1
                    formula = f"=({get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(rate_col_idx)}{current_row})+{get_column_letter(labour_amount_col_idx)}{current_row}"
                    row_values.append(formula)
                elif key == "to_be_maintained":
                    qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                    rate_col_idx = COLUMN_HEADERS.index("Rate in Rs") + 1
                    formula = f"={get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(rate_col_idx)}{current_row}"
                    row_values.append(formula)
                else:
                    row_values.append(item.get(key))
            worksheet.append(row_values)
            # Apply border to data rows
            for col_idx in range(1, len(COLUMN_HEADERS) + 1):
                worksheet.cell(row=current_row, column=col_idx).border = thin_border
            current_row += 1
        data_end_row = current_row - 1

        # Write summary section
        # Sub Total
        sub_total_row = [None] * len(COLUMN_HEADERS)
        sub_total_desc_col_idx = COLUMN_HEADERS.index("Description")
        sub_total_row[sub_total_desc_col_idx] = "Sub Total"
        total_in_rs_col_idx = COLUMN_HEADERS.index("Total in Rs") + 1
        sub_total_formula = f"=SUM({get_column_letter(total_in_rs_col_idx)}{data_start_row}:{get_column_letter(total_in_rs_col_idx)}{data_end_row})"
        
        worksheet.append(sub_total_row)
        sub_total_desc_cell = worksheet.cell(row=current_row, column=sub_total_desc_col_idx + 1)
        sub_total_desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        sub_total_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
        sub_total_cell.value = sub_total_formula
        sub_total_cell.data_type = 'f'
        sub_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        for col_idx in range(1, len(COLUMN_HEADERS) + 1):
            worksheet.cell(row=current_row, column=col_idx).border = thin_border
        sub_total_row_idx = current_row
        current_row += 1

        # GST @18%
        gst_row = [None] * len(COLUMN_HEADERS)
        gst_desc_col_idx = COLUMN_HEADERS.index("Description")
        gst_row[gst_desc_col_idx] = f"GST @{int(GST_RATE*100)}%"
        gst_formula = f"={get_column_letter(total_in_rs_col_idx)}{sub_total_row_idx}*{GST_RATE}"
        
        worksheet.append(gst_row)
        gst_desc_cell = worksheet.cell(row=current_row, column=gst_desc_col_idx + 1)
        gst_desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        gst_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
        gst_cell.value = gst_formula
        gst_cell.data_type = 'f'
        gst_cell.alignment = Alignment(horizontal='right', vertical='center')
        for col_idx in range(1, len(COLUMN_HEADERS) + 1):
            worksheet.cell(row=current_row, column=col_idx).border = thin_border
        gst_row_idx = current_row
        current_row += 1

        # Grand Total (All Inclusive)
        grand_total_row = [None] * len(COLUMN_HEADERS)
        grand_total_desc_col_idx = COLUMN_HEADERS.index("Description")
        grand_total_row[grand_total_desc_col_idx] = "Grand Total (All Inclusive)"
        grand_total_formula = f"={get_column_letter(total_in_rs_col_idx)}{sub_total_row_idx}+{get_column_letter(total_in_rs_col_idx)}{gst_row_idx}"
        
        worksheet.append(grand_total_row)
        grand_total_desc_cell = worksheet.cell(row=current_row, column=grand_total_desc_col_idx + 1)
        grand_total_desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        grand_total_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
        grand_total_cell.value = grand_total_formula
        grand_total_cell.data_type = 'f'
        grand_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        for col_idx in range(1, len(COLUMN_HEADERS) + 1):
            worksheet.cell(row=current_row, column=col_idx).border = thin_border
        grand_total_row_idx = current_row
        current_row += 1

        # Add one blank row after Grand Total
        current_row += 1

        # Signature Block
        worksheet.cell(row=current_row, column=1, value="PLACE : KALYAN")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="DATE : 05-07-2025")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Allocation : 47000 070 443 32")
        current_row += 5 # Empty Space: Leave 4-5 blank rows for buffer/visual clarity.

        # Signature Row: Create four signature placeholders aligned horizontally across the page
        bold_center_alignment = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        # SSE/WKS
        cell = worksheet.cell(row=current_row, column=1, value="SSE/WKS")
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # SSE/MW
        cell = worksheet.cell(row=current_row, column=3, value="SSE/MW")
        worksheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=5, value="DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Sr.DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=7, value="Sr.DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Signature Block
        worksheet.cell(row=current_row, column=1, value="PLACE : KALYAN")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="DATE : 05-07-2025")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Allocation : 47000 070 443 32")
        current_row += 5 # Empty Space: Leave 4-5 blank rows for buffer/visual clarity.

        # Signature Row: Create four signature placeholders aligned horizontally across the page
        bold_center_alignment = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        # SSE/WKS
        cell = worksheet.cell(row=current_row, column=1, value="SSE/WKS")
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # SSE/MW
        cell = worksheet.cell(row=current_row, column=3, value="SSE/MW")
        worksheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=5, value="DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Sr.DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=7, value="Sr.DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Signature Block
        worksheet.cell(row=current_row, column=1, value="PLACE : KALYAN")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="DATE : 05-07-2025")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Allocation : 47000 070 443 32")
        current_row += 5 # Empty Space: Leave 4-5 blank rows for buffer/visual clarity.

        # Signature Row: Create four signature placeholders aligned horizontally across the page
        bold_center_alignment = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        # SSE/WKS
        cell = worksheet.cell(row=current_row, column=1, value="SSE/WKS")
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # SSE/MW
        cell = worksheet.cell(row=current_row, column=3, value="SSE/MW")
        worksheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=5, value="DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Sr.DEE (TRS) KALYAN
        cell = worksheet.cell(row=current_row, column=7, value="Sr.DEE (TRS) KALYAN")
        worksheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
        cell.alignment = bold_center_alignment
        cell.font = bold_font

        # Auto-fit columns
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(output_path)
        return True, "Estimate exported successfully"
    except Exception as e:
        return False, f"Error exporting estimate: {str(e)}"