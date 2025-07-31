# features/estimates/header_writer.py

from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from .constants import COLUMN_HEADERS

def write_estimate_header_block(worksheet, estimate_no="[ESTIMATE_NO]"):
    """
    Writes the header block at the top of the estimate report spanning the full table width.
    
    Args:
        worksheet: The openpyxl worksheet object
        estimate_no: The estimate number to display
    
    Returns:
        int: The next available row after the header block
    """
    # Get the total number of columns in the estimate table
    total_columns = len(COLUMN_HEADERS)  # Should be 10 columns
    
    # Define styles
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define border style (thin black border)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # First row: Estimate No. on left, ELECTRICAL DEPARTMENT on right
    # Left side - Estimate No. (spans first half of columns)
    estimate_cell = worksheet.cell(row=current_row, column=1, value=f"ESTIMATE NO. : {estimate_no}")
    estimate_cell.font = bold_font
    estimate_cell.alignment = left_alignment
    estimate_cell.border = thin_border
    
    # Merge columns 1 to half of total columns for estimate number
    mid_column = total_columns // 2
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=mid_column)
    
    # Right side - ELECTRICAL DEPARTMENT (spans second half of columns)
    dept_cell = worksheet.cell(row=current_row, column=mid_column + 1, value="ELECTRICAL DEPARTMENT")
    dept_cell.font = bold_font
    dept_cell.alignment = right_alignment
    dept_cell.border = thin_border
    
    # Merge remaining columns for department
    worksheet.merge_cells(start_row=current_row, start_column=mid_column + 1, end_row=current_row, end_column=total_columns)
    
    current_row += 1
    
    # Second row: Empty on left, DIVISION : MUMBAI on right
    # Left side - Empty (spans first half)
    empty_cell = worksheet.cell(row=current_row, column=1, value="")
    empty_cell.border = thin_border
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=mid_column)
    
    # Right side - DIVISION : MUMBAI (spans second half)
    division_cell = worksheet.cell(row=current_row, column=mid_column + 1, value="DIVISION : MUMBAI")
    division_cell.font = normal_font
    division_cell.alignment = right_alignment
    division_cell.border = thin_border
    worksheet.merge_cells(start_row=current_row, start_column=mid_column + 1, end_row=current_row, end_column=total_columns)
    
    current_row += 1
    
    # Third row: Empty on left, PLACE OF WORK : ELECTRIC LOCO SHED, KALYAN on right
    # Left side - Empty (spans first half)
    empty_cell2 = worksheet.cell(row=current_row, column=1, value="")
    empty_cell2.border = thin_border
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=mid_column)
    
    # Right side - PLACE OF WORK (spans second half)
    place_cell = worksheet.cell(row=current_row, column=mid_column + 1, value="PLACE OF WORK : ELECTRIC LOCO SHED, KALYAN")
    place_cell.font = normal_font
    place_cell.alignment = right_alignment
    place_cell.border = thin_border
    worksheet.merge_cells(start_row=current_row, start_column=mid_column + 1, end_row=current_row, end_column=total_columns)
    
    current_row += 1
    
    # Add one blank row after the header block
    current_row += 1
    
    return current_row

def write_estimate_header_block_simple(worksheet, estimate_no="[ESTIMATE_NO]"):
    """
    Writes a simple header block without borders at the top of the estimate report spanning the full table width.
    
    Args:
        worksheet: The openpyxl worksheet object
        estimate_no: The estimate number to display
    
    Returns:
        int: The next available row after the header block
    """
    # Get the total number of columns in the estimate table
    total_columns = len(COLUMN_HEADERS)  # Should be 10 columns
    
    # Define styles
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    current_row = 1
    
    # First row: Estimate No. on left, ELECTRICAL DEPARTMENT on right
    # Left side - Estimate No.
    estimate_cell = worksheet.cell(row=current_row, column=1, value=f"ESTIMATE NO. : {estimate_no}")
    estimate_cell.font = bold_font
    estimate_cell.alignment = left_alignment
    
    # Right side - ELECTRICAL DEPARTMENT (positioned at the end of the table)
    dept_cell = worksheet.cell(row=current_row, column=total_columns, value="ELECTRICAL DEPARTMENT")
    dept_cell.font = bold_font
    dept_cell.alignment = right_alignment
    
    current_row += 1
    
    # Second row: Empty on left, DIVISION : MUMBAI on right
    # Right side - DIVISION : MUMBAI
    division_cell = worksheet.cell(row=current_row, column=total_columns, value="DIVISION : MUMBAI")
    division_cell.font = normal_font
    division_cell.alignment = right_alignment
    
    current_row += 1
    
    # Third row: Empty on left, PLACE OF WORK : ELECTRIC LOCO SHED, KALYAN on right
    # Right side - PLACE OF WORK
    place_cell = worksheet.cell(row=current_row, column=total_columns, value="PLACE OF WORK : ELECTRIC LOCO SHED, KALYAN")
    place_cell.font = normal_font
    place_cell.alignment = right_alignment
    
    current_row += 1
    
    # Add one blank row after the header block
    current_row += 1
    
    return current_row
