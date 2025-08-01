# features/estimates/formatter.py

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .constants import COLUMN_STYLES
from .utils import autosize_columns


def apply_header_style(worksheet, start_row, end_row, start_col, end_col):
    """
    Applies styling to the header row: bold font, borders.
    """
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    from openpyxl.styles import PatternFill
    grey_fill = PatternFill(
        start_color="FFD9D9D9",
        end_color="FFD9D9D9",
        fill_type="solid")

    for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col):
        for cell in row:
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = grey_fill


def apply_data_row_style(worksheet, row_index, start_col, end_col):
    """
    Applies basic styling to data rows.
    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    from .constants import COLUMN_HEADERS
    
    for col_idx in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row_index, column=col_idx)
        cell.border = thin_border
        
        # Check if this is the Description column for left alignment
        if col_idx <= len(COLUMN_HEADERS):
            column_name = COLUMN_HEADERS[col_idx - 1]  # Convert to 0-based index
            if column_name == "Description":
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_summary_style(worksheet, start_row, end_row, start_col, end_col):
    """
    Applies styling to summary rows: bold, borders, shaded background.
    """
    summary_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Light grey fill
    from openpyxl.styles import PatternFill
    grey_fill = PatternFill(
        start_color="FFD9D9D9",
        end_color="FFD9D9D9",
        fill_type="solid")

    for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col):
        for cell in row:
            cell.font = summary_font
            cell.border = thin_border
            cell.fill = grey_fill


def format_columns(
        worksheet,
        data_start_row,
        data_end_row,
        column_headers,
        summary_start_row=None,
        summary_end_row=None):
    """
    Applies number formatting to specified columns for both data and summary rows.
    """
    for col_name, style_info in COLUMN_STYLES.items():
        try:
            # +1 because openpyxl is 1-indexed
            col_idx = column_headers.index(col_name) + 1
            number_format = style_info.get('format')
            if number_format:
                # Format data rows
                for row_idx in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = number_format

                # Format summary rows if provided
                if summary_start_row is not None and summary_end_row is not None:
                    for row_idx in range(
                            summary_start_row, summary_end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = number_format
        except ValueError:
            # Column not found, ignore
            pass


def style_work_name_row(worksheet, work_name_row_idx, total_columns):
    """
    Applies specific styling to the work name row: height, bold font, center alignment.

    Note: Background color fill was removed for simplified styling.
    """
    # Set row height
    worksheet.row_dimensions[work_name_row_idx].height = 30

    # Define styles
    work_name_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    # Optional light blue fill color for work name row
    work_name_fill = PatternFill(
        start_color="FFE6F3FF",
        end_color="FFE6F3FF",
        fill_type="solid")

    # Apply styles to all cells in the work name row
    for col_idx in range(1, total_columns + 1):
        cell = worksheet.cell(row=work_name_row_idx, column=col_idx)
        cell.font = work_name_font
        cell.alignment = center_alignment


def apply_all_styles_and_formats(
        worksheet,
        header_row_idx,
        data_start_row,
        data_end_row,
        summary_start_row,
        summary_end_row,
        column_headers,
        all_data,
        work_name_row_idx):
    """
    Applies all necessary styles and formats to the worksheet.
    """
    # Apply work name row style (early so later formats don't override it)
    style_work_name_row(worksheet, work_name_row_idx, len(column_headers))

    # Apply header style
    apply_header_style(
        worksheet,
        header_row_idx,
        header_row_idx,
        1,
        len(column_headers))

    # Apply data row styles
    for row_idx in range(data_start_row, data_end_row + 1):
        apply_data_row_style(worksheet, row_idx, 1, len(column_headers))

    # Apply summary style
    apply_summary_style(
        worksheet,
        summary_start_row,
        summary_end_row,
        1,
        len(column_headers))

    # Format columns (e.g., currency) - include summary rows
    format_columns(
        worksheet,
        data_start_row,
        data_end_row,
        column_headers,
        summary_start_row,
        summary_end_row)

    # Autosize columns
    autosize_columns(worksheet, all_data)
