# features/estimates/utils.py

from openpyxl.utils import get_column_letter

def get_cell_reference(row, col):
    """
    Returns the Excel cell reference (e.g., 'A1') for a given row and column.
    """
    return f"{get_column_letter(col)}{row}"

def autosize_columns(worksheet, data):
    """
    Autosizes columns based on the content in the worksheet.
    """
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column index
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # Add a little extra space
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

def generate_formula(row, col_qty, col_rate, col_labour_amount=None):
    """
    Generates Excel formulas for Labour Amount and Total in Rs.
    """
    qty_ref = get_cell_reference(row, col_qty)
    rate_ref = get_cell_reference(row, col_rate)

    if col_labour_amount:
        # Formula for Labour Amount: Qty * Labour rate
        labour_rate_ref = get_cell_reference(row, col_rate) # Assuming col_rate is labour_rate for this formula
        return f"={qty_ref}*{labour_rate_ref}"
    else:
        # Formula for Total in Rs: (Qty * Rate) + Labour Amount
        labour_amount_ref = get_cell_reference(row, col_labour_amount)
        return f"=({qty_ref}*{rate_ref})+{labour_amount_ref}"
