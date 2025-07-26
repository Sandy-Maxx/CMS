# features/estimates/writer.py

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from .constants import COLUMN_HEADERS, COLUMN_MAPPING, GST_RATE
from .utils import get_cell_reference

def write_header(worksheet, work_details):
    """
    Writes the column headers to the worksheet, starting from A1.
    """
    worksheet.append(COLUMN_HEADERS)
    return 1 # Returns the row index of the header (which is now 1)

def write_work_description_row(worksheet, data_row, current_row):
    """
    Writes the special 'A' row for work description.
    """
    row_values = []
    for header in COLUMN_HEADERS:
        key = COLUMN_MAPPING.get(header)
        row_values.append(data_row.get(key))
    worksheet.append(row_values)
    return current_row + 1

def write_schedule_data_rows(worksheet, data, start_row):
    """
    Writes schedule data rows with Excel formulas.
    """
    current_row = start_row
    for item in data:
        row_values = []
        for header in COLUMN_HEADERS:
            key = COLUMN_MAPPING.get(header)
            if key == "labour_amount":
                # Qty column is 3rd (index 2), Labour rate in Rs. is 7th (index 6)
                qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                labour_rate_col_idx = COLUMN_HEADERS.index("Labour rate in Rs.") + 1
                formula = f"={get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(labour_rate_col_idx)}{current_row}"
                row_values.append(formula)
            elif key == "total_in_rs":
                # Qty column is 3rd (index 2), Rate in Rs is 4th (index 3), Labour Amount is 8th (index 7)
                qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                rate_col_idx = COLUMN_HEADERS.index("Rate in Rs") + 1
                labour_amount_col_idx = COLUMN_HEADERS.index("Labour Amount") + 1
                formula = f"=({get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(rate_col_idx)}{current_row})+{get_column_letter(labour_amount_col_idx)}{current_row}"
                row_values.append(formula)
            elif key == "to_be_maintained":
                qty_col_idx = COLUMN_HEADERS.index("Qty") + 1
                rate_col_idx = COLUMN_HEADERS.index("Rate in Rs") + 1
                formula = f"={get_column_letter(qty_col_idx)}{current_row}*{get_column_letter(rate_col_idx)}{current_row}"
                row_values.append(formula)
            else:
                row_values.append(item.get(key))
        worksheet.append(row_values)
        current_row += 1
    return current_row

def write_summary_section(worksheet, data_start_row, data_end_row):
    """
    Writes the summary section with formulas: Sub Total, GST, Grand Total.
    """
    current_row = data_end_row + 1

    # Sub Total
    sub_total_row = [None] * len(COLUMN_HEADERS)
    sub_total_desc_col_idx = COLUMN_HEADERS.index("Description")
    sub_total_row[sub_total_desc_col_idx] = "Sub Total"
    total_in_rs_col_idx = COLUMN_HEADERS.index("Total in Rs") + 1
    sub_total_formula = f"=SUM({get_column_letter(total_in_rs_col_idx)}{data_start_row}:{get_column_letter(total_in_rs_col_idx)}{data_end_row})"
    
    # Append the row and then set the formula cell type and alignment
    worksheet.append(sub_total_row)
    sub_total_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
    sub_total_cell.value = sub_total_formula
    sub_total_cell.data_type = 'f'

    sub_total_row_idx = current_row
    current_row += 1

    # GST @18%
    gst_row = [None] * len(COLUMN_HEADERS)
    gst_desc_col_idx = COLUMN_HEADERS.index("Description")
    gst_row[gst_desc_col_idx] = f"GST @{int(GST_RATE*100)}%"
    gst_formula = f"={get_column_letter(total_in_rs_col_idx)}{sub_total_row_idx}*{GST_RATE}"
    
    # Append the row and then set the formula cell type and alignment
    worksheet.append(gst_row)
    gst_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
    gst_cell.value = gst_formula
    gst_cell.data_type = 'f'

    gst_row_idx = current_row
    current_row += 1

    # Grand Total (All Inclusive)
    grand_total_row = [None] * len(COLUMN_HEADERS)
    grand_total_desc_col_idx = COLUMN_HEADERS.index("Description")
    grand_total_row[grand_total_desc_col_idx] = "Grand Total (All Inclusive)"
    grand_total_formula = f"={get_column_letter(total_in_rs_col_idx)}{sub_total_row_idx}+{get_column_letter(total_in_rs_col_idx)}{gst_row_idx}"
    
    # Append the row and then set the formula cell type and alignment
    worksheet.append(grand_total_row)
    grand_total_cell = worksheet.cell(row=current_row, column=total_in_rs_col_idx)
    grand_total_cell.value = grand_total_formula
    grand_total_cell.data_type = 'f'

    grand_total_row_idx = current_row
    current_row += 1

    # Add one blank row after Grand Total
    current_row += 1

    # Signature Block

    # Header Rows (Top-left aligned)
    worksheet.cell(row=current_row, column=1, value="PLACE : KALYAN")
    current_row += 1
    worksheet.cell(row=current_row, column=1, value="DATE : 05-07-2025")
    current_row += 1
    worksheet.cell(row=current_row, column=1, value="Allocation : 47000 070 443 32")
    current_row += 1

    # Empty Space: Leave 4-5 blank rows for buffer/visual clarity.
    current_row += 5

    # Signature Row: Create four signature placeholders aligned horizontally across the page
    bold_center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    # SSE/WKS
    cell = worksheet.cell(row=current_row, column=1, value="SSE/WKS")
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    cell.alignment = bold_center_alignment
    cell.font = bold_font

    # SSE/MW
    cell = worksheet.cell(row=current_row, column=3, value="SSE/MW")
    worksheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    cell.alignment = bold_center_alignment
    cell.font = bold_font

    # DEE (TRS) KALYAN
    cell = worksheet.cell(row=current_row, column=5, value="DEE (TRS) KALYAN")
    worksheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
    cell.alignment = bold_center_alignment
    cell.font = bold_font

    # Sr.DEE (TRS) KALYAN
    cell = worksheet.cell(row=current_row, column=7, value="Sr.DEE (TRS) KALYAN")
    worksheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
    cell.alignment = bold_center_alignment
    cell.font = bold_font

    current_row += 1 # Move to the next row after signature block

    return sub_total_row_idx, grand_total_row_idx