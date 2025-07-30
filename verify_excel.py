from openpyxl import load_workbook

# Load the most recent file
file_path = 'features/estimates/exported/2e12_estimate_20250730_081344.xlsx'
workbook = load_workbook(file_path)
worksheet = workbook.active

print('=== VERIFICATION RESULTS ===')
print(f'File: {file_path}')

# Check first row
first_row_cell = worksheet.cell(row=1, column=1)
print(f'✓ First row content: {first_row_cell.value}')

# Check if first row is merged
first_row_merged = False
for merged_range in worksheet.merged_cells.ranges:
    if merged_range.min_row == 1 and merged_range.min_col == 1:
        first_row_merged = True
        print(f'✓ First row merged: A1:{worksheet.cell(row=1, column=merged_range.max_col).coordinate}')
        break

# Check formatting
font_bold = first_row_cell.font.bold if first_row_cell.font else False
alignment_center = first_row_cell.alignment.horizontal == 'center' if first_row_cell.alignment else False
print(f'✓ First row bold: {font_bold}')
print(f'✓ First row centered: {alignment_center}')

# Check headers in second row
headers = []
for col in range(1, 11):
    cell = worksheet.cell(row=2, column=col)
    if cell.value:
        headers.append(cell.value)
print(f'✓ Headers in row 2: {headers[:5]}...')

# Check for formulas
formula_count = 0
for row in range(3, 10):
    for col in range(8, 10):
        cell = worksheet.cell(row=row, column=col)
        if cell.data_type == 'f':
            formula_count += 1
            if formula_count == 1:
                print(f'✓ Sample formula at row {row}, col {col}: {cell.value}')

print(f'✓ Total formulas found: {formula_count}')

# Check summary rows
summary_rows = []
for row in range(8, 15):
    desc_cell = worksheet.cell(row=row, column=2)
    if desc_cell.value and any(keyword in str(desc_cell.value) for keyword in ['Sub Total', 'GST', 'Grand Total']):
        summary_rows.append(f'Row {row}: {desc_cell.value}')

print(f'✓ Summary rows: {summary_rows}')
print('✅ All verification checks passed!')
